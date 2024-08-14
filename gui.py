
from pathlib import Path
from tkinter import Tk, Canvas, Entry, Button, PhotoImage, messagebox
import re
import openpyxl

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path(r"C:\Users\Siva\Desktop\build\assets\frame0")

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)

def validate_email(email):
    # Regex pattern for validating an Email
    pattern = r'^[a-zA-Z0-9._%-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email)

def validate_fields(fields):
    for field in fields:
        if not field.get():
            return False
    return True

def save_to_excel(data):
    file_path = 'submissions.xlsx'
    
    if not Path(file_path).exists():
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Submissions'
        headers = ["ID", "First Name", "Last Name", "Email", "Phone", "Company", "Designation"]
        sheet.append(headers)
        workbook.save(file_path)

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    new_id = sheet.max_row
    data.insert(0, new_id)
    sheet.append(data)
    workbook.save(file_path)

def submit():
    fields = [entry_1, entry_3, entry_5, entry_2, entry_4, entry_6]
    if not validate_fields(fields):
        messagebox.showerror("Error", "All fields must be filled.")
        return
    
    email = entry_5.get()
    if not validate_email(email):
        messagebox.showerror("Error", "Invalid email format.")
        return

    data = [field.get() for field in fields]
    save_to_excel(data)
    
    for field in fields:
        field.delete(0, 'end')
    
    messagebox.showinfo("Success", "Submission saved successfully.")

window = Tk()

window.geometry("650x350")
window.configure(bg = "#FFFFFF")

canvas = Canvas(
    window,
    bg = "#FFFFFF",
    height = 350,
    width = 650,
    bd = 0,
    highlightthickness=0,
    relief = "ridge"
)

canvas.place(x = 0, y = 0)
canvas.create_rectangle(
    0.0,
    63.0,
    650.0,
    353.0,
    fill="#ADD8E6",
    outline="")

canvas.create_rectangle(
    2.0,
    0.0,
    652.0,
    60.0,
    fill="#FFFFFF",
    outline="")

canvas.create_text(
    235.0,
    9.0,
    anchor="nw",
    text="Contact Book",
    fill="#000000",
    font=("BadScript Regular", 40 * -1)
)

canvas.create_text(
    8.0,
    85.0,
    anchor="nw",
    text="First name:",
    fill="#000000",
    font=("Inter Medium", 20 * -1)
)

canvas.create_text(
    352.0,
    148.0,
    anchor="nw",
    text="Company:",
    fill="#000000",
    font=("Inter Medium", 20 * -1)
)

canvas.create_text(
    327.0,
    211.0,
    anchor="nw",
    text="Designation:",
    fill="#000000",
    font=("Inter Medium", 20 * -1)
)

canvas.create_text(
    347.0,
    85.0,
    anchor="nw",
    text="Phone no.:",
    fill="#000000",
    font=("Inter Medium", 20 * -1)
)

entry_image_1 = PhotoImage(
    file=relative_to_assets("entry_1.png"))
entry_bg_1 = canvas.create_image(
    208.0,
    97.5,
    image=entry_image_1
)
entry_1 = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#000716",
    highlightthickness=0
)
entry_1.place(
    x=139.0,
    y=82.0,
    width=138.0,
    height=29.0
)

entry_image_2 = PhotoImage(
    file=relative_to_assets("entry_2.png"))
entry_bg_2 = canvas.create_image(
    543.0,
    97.5,
    image=entry_image_2
)
entry_2 = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#000716",
    highlightthickness=0
)
entry_2.place(
    x=474.0,
    y=82.0,
    width=138.0,
    height=29.0
)

entry_image_3 = PhotoImage(
    file=relative_to_assets("entry_3.png"))
entry_bg_3 = canvas.create_image(
    208.0,
    159.5,
    image=entry_image_3
)
entry_3 = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#000716",
    highlightthickness=0
)
entry_3.place(
    x=139.0,
    y=144.0,
    width=138.0,
    height=29.0
)

entry_image_4 = PhotoImage(
    file=relative_to_assets("entry_4.png"))
entry_bg_4 = canvas.create_image(
    543.0,
    159.5,
    image=entry_image_4
)
entry_4 = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#000716",
    highlightthickness=0
)
entry_4.place(
    x=474.0,
    y=144.0,
    width=138.0,
    height=29.0
)

entry_image_5 = PhotoImage(
    file=relative_to_assets("entry_5.png"))
entry_bg_5 = canvas.create_image(
    208.0,
    223.5,
    image=entry_image_5
)
entry_5 = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#000716",
    highlightthickness=0
)
entry_5.place(
    x=139.0,
    y=208.0,
    width=138.0,
    height=29.0
)

entry_image_6 = PhotoImage(
    file=relative_to_assets("entry_6.png"))
entry_bg_6 = canvas.create_image(
    543.0,
    224.5,
    image=entry_image_6
)
entry_6 = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#000716",
    highlightthickness=0
)
entry_6.place(
    x=474.0,
    y=209.0,
    width=138.0,
    height=29.0
)

canvas.create_text(
    5.0,
    147.0,
    anchor="nw",
    text="Last name:",
    fill="#1E1E1E",
    font=("Inter Medium", 20 * -1)
)

canvas.create_text(
    40.0,
    211.0,
    anchor="nw",
    text="E-mail:",
    fill="#000000",
    font=("Inter Medium", 20 * -1)
)

button_image_1 = PhotoImage(
    file=relative_to_assets("button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=submit,
    relief="flat"
)
button_1.place(
    x=208.0,
    y=266.0,
    width=212.0,
    height=56.0
)

image_image_1 = PhotoImage(
    file=relative_to_assets("image_1.png"))
image_1 = canvas.create_image(
    180.0,
    30.0,
    image=image_image_1
)
window.resizable(False, False)
window.mainloop()
